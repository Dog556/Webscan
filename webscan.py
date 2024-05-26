import os
import requests
from bs4 import BeautifulSoup
import concurrent.futures
from pyfiglet import Figlet
from datetime import datetime
from openpyxl import Workbook
from tqdm import tqdm
import argparse
import socket
import re

#参数设置
parser = argparse.ArgumentParser(description="Process a file.")
parser.add_argument("-f", "--file", type=str, required=True, help="指定一个txt文档")
parser.add_argument("-t", "--thread", type=str, required=False, help="线程最大为10默认")
parser.add_argument("-c", "--code", type=str, required=False, help="指定指定状态码，默认为200")
args = parser.parse_args()
filepath = args.file
thread = args.thread
code = args.code
if code == None:
    code = 200
if thread==None:
    thread = 10
thread = int(thread)
port = [80,81,443,1433,3306,1521,5432]
keywords = ['password', 'username','login','登录']
rows_to_write = []
add_url=[]
add_title=[]
add_time=[]
add_mid=[]
add_ip=[]
add_login=[]
fs = Figlet(font='slant')
print(fs.renderText('WebScan'))
print("\033[0;31;40mv1.0by the Carrot\033[0m")
headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36 Edg/118.0.2088.46',
         'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6'
}

now_time = datetime.now()
now_time_str = now_time.strftime("%Y-%m-%d-%H-%M-%S")
file_path=now_time_str + '.xlsx'
wb = Workbook()
ws = wb.active
ws['A1'] = '时间'
ws['B1'] = '标题'
ws['C1'] = 'url'
ws['D1'] = '中间件'
ws['E1'] = 'IP'
ws['F1'] = '是否有登录框'
ws['G1'] = '状态码'
ws.column_dimensions['C'].width = 34
ws.column_dimensions['B'].width = 40
ws.column_dimensions['A'].width = 22
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 20

#判断是否有登录
def check_keywords_in_text(text, keywords):
    pattern = r'\b(' + '|'.join(re.escape(keyword) for keyword in keywords) + r')\b'
    if re.search(pattern, text, re.IGNORECASE):
        return True
    else:
        return False

#正则匹配是否为IP
def is_valid_domain(domain):
    pattern = re.compile(
        r'^(?:(?:[A-Z0-9-]{1,63}\.){1,125}(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|'
        r'localhost|\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})$',
        re.IGNORECASE)
    return bool(pattern.match(domain))


def get_ip_from_domain(domain):
    if is_valid_domain(domain):
        try:
            ip_address = socket.gethostbyname(domain)
            return ip_address
        except socket.gaierror:
            return None  # 域名不存在或其他网络错误
    else:
        return None  # 不是一个有效的域名

def remove_http_prefix(url):
    # 匹配 http:// 或 https:// 开头的部分，并替换为空字符串
    return re.sub(r'^https?://', '', url)

#获取标题
def get_title(url,headers):
    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.content, 'html.parser')
    # text_content = r.text
    # text_content = r.content.decode('utf-8')
    if r.status_code == 200:
        title = soup.find('title').get_text()
        title = str(title)
        add_title.append(title)
        return title

#读取文件
def txtscan(filename):
    with open(filename,'r', encoding='utf-8') as file:
        for line in file:
            mup = file.read().splitlines()
            return mup

#写入进度条
def preprocess_data(index, url, title, time ,mid):
    row_data = (url, title, time, "200",mid)
    rows_to_write.append(row_data)
#主程序，判断code 返回数组
def process_url(url,code):
    try:
        r = requests.get(url,headers=headers,timeout=2)
        title = get_title(url,headers)
        if r.status_code == code:
            server_header = r.headers.get('Server')
            domain = remove_http_prefix(url).rstrip('/') #判断是否为域名 如果不是 就是返回空
            ip = get_ip_from_domain(domain)
            if check_keywords_in_text(r.text,keywords):#登录判断
                add_login.append("yes")
            else:
                add_login.append("no")
            if server_header == None:
                server_header = 'None'
            now_time1 = datetime.now()
            print('\033[92m' + f"{url},{title},{r.status_code}"+ "\033[0m")
            add_url.append(url)
            add_time.append(now_time1)
            add_mid.append(server_header)
            add_ip.append(ip)
    except requests.exceptions.Timeout:
        print("\033[0;37;41m网站异常,Timeout！！\033[0m",f"\033[0;31;40m{url}\033[0m")
    except requests.exceptions.RequestException as e:
        print('\033[91m' + f"{url} error！" + '\033[0m')

#线程设置
with concurrent.futures.ThreadPoolExecutor(max_workers=thread) as executor:
    urls = txtscan(filepath)
    # 提交任务到线程池
    futures = [executor.submit(process_url,url,code) for url in urls]

    for future in concurrent.futures.as_completed(futures):
        pass

print("数据正在生成中.....")


# 使用ThreadPoolExecutor来并行处理数据
with concurrent.futures.ThreadPoolExecutor() as executor, tqdm(total=len(add_url)) as pbar:
    for index, (url, title, time, mid) in enumerate(zip(add_url, add_title, add_time ,add_mid), start=2):
        executor.submit(preprocess_data, index, url, title, time, mid)
        pbar.update()  # 更新进度条

for row_num, (url, title, time, mid , ip,login) in enumerate(zip(add_url, add_title, add_time,add_mid,add_ip,add_login), 2):
    ws.cell(row=row_num, column=3, value=url)     # URL写入第3列
    ws.cell(row=row_num, column=2, value=title)   # 标题写入第2列
    ws.cell(row=row_num, column=1, value=time)    # 时间写入第1列
    ws.cell(row=row_num, column=4, value=mid)
    ws.cell(row=row_num, column=5, value=ip)
    ws.cell(row=row_num, column=6, value=login)
    ws.cell(row=row_num, column=7, value="200")   # 固定值"200"写入第4列 上面已经判断过code等于200

wb.save(f'{now_time_str}.xlsx')
os.startfile(file_path)